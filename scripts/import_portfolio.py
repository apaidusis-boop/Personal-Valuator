"""Importador de carteira real — BR (XP xlsx) + US (JPM csv).

Popula `portfolio_positions` com quantidade, preço médio (cost basis) e data
de carregamento em ambas as DBs. Substitui `seed_portfolio.py` para os
holdings equities reais (tesouro/debêntures/CRAs ficam fora — o schema
actual só modela equity).

Formato BR: XP "PosiçãoDetalhada" xlsx — secções "Ações", "Fundos Imobiliários".
Formato US: J.P. Morgan Wealth "positions.csv".

Uso:
    python scripts/import_portfolio.py --br "<path>.xlsx" --us "<path>.csv"
    python scripts/import_portfolio.py --br "<path>.xlsx"      # só BR
    python scripts/import_portfolio.py --dry-run --br "..."    # não grava
"""
from __future__ import annotations

import argparse
import csv
import re
import sqlite3
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

DB_BR = ROOT / "data" / "br_investments.db"
DB_US = ROOT / "data" / "us_investments.db"

TODAY = date.today().isoformat()


def ensure_schema(conn: sqlite3.Connection) -> None:
    cols = {r[1] for r in conn.execute("PRAGMA table_info(portfolio_positions)").fetchall()}
    if "quantity" not in cols:
        conn.execute("ALTER TABLE portfolio_positions ADD COLUMN quantity REAL")
    if "notes" not in cols:
        conn.execute("ALTER TABLE portfolio_positions ADD COLUMN notes TEXT")
    # fixed_income_positions (idempotente)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fixed_income_positions (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            name              TEXT NOT NULL,
            kind              TEXT NOT NULL,
            indexador         TEXT,
            spread_taxa       REAL,
            cdi_pct           REAL,
            entry_date        TEXT,
            maturity_date     TEXT,
            quantity          REAL,
            entry_unit_price  REAL,
            valor_aplicado    REAL,
            valor_atual       REAL NOT NULL,
            currency          TEXT NOT NULL DEFAULT 'BRL',
            source            TEXT,
            fetched_at        TEXT NOT NULL,
            notes             TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_fi_maturity "
                 "ON fixed_income_positions(maturity_date)")


def _parse_brl(v: object) -> float | None:
    """Parser para números no formato BR (XP): '.' é milhar, ',' é decimal.
    'R$ 1.234,56' → 1234.56
    '1.026'        → 1026  (sem vírgula, '.' vira milhar)
    '120,27'       → 120.27
    Para números sem separadores (US format), retorna como está."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace("\xa0", " ").strip()
    s = s.replace("%", "").strip()
    if "," in s:
        # BR: decimal = ",". Remove "." milhares.
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        # sem vírgula: "." é milhar (XP). Cuidado: pode ser decimal US.
        # Heurística: se depois do último "." há exactamente 3 dígitos e inteiro
        # antes > 0, é milhar. Caso contrário decimal.
        parts = s.split(".")
        if len(parts) >= 2 and len(parts[-1]) == 3 and all(p.isdigit() for p in parts):
            s = "".join(parts)
    try:
        return float(s)
    except ValueError:
        return None


def _parse_us_number(v: object) -> float | None:
    """US format: ',' é milhar, '.' é decimal. '1,234.56' → 1234.56."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _upsert_company(conn: sqlite3.Connection, ticker: str, name: str,
                    currency: str, sector: str | None = None) -> None:
    conn.execute(
        """INSERT INTO companies (ticker, name, sector, is_holding, currency)
           VALUES (?, ?, ?, 1, ?)
           ON CONFLICT(ticker) DO UPDATE SET
             name=COALESCE(excluded.name, companies.name),
             sector=COALESCE(excluded.sector, companies.sector),
             is_holding=1""",
        (ticker, name, sector, currency),
    )


def _upsert_position(conn: sqlite3.Connection, ticker: str, quantity: float,
                     entry_price: float, entry_date: str,
                     notes: str | None = None) -> None:
    # weight recalculado no fim (pass final); por enquanto placeholder
    conn.execute(
        """INSERT INTO portfolio_positions
             (ticker, weight, entry_date, entry_price, active, quantity, notes)
           VALUES (?, 0.0, ?, ?, 1, ?, ?)
           ON CONFLICT(ticker, entry_date) DO UPDATE SET
             entry_price=excluded.entry_price,
             quantity=excluded.quantity,
             active=1,
             notes=excluded.notes""",
        (ticker, entry_date, entry_price, quantity, notes),
    )


def _recompute_weights(conn: sqlite3.Connection) -> None:
    """Pesos = valor actual (quantity × último close) / total. Se sem preço,
    cai para cost basis (qty × entry_price)."""
    rows = conn.execute(
        "SELECT ticker, entry_date, quantity, entry_price FROM portfolio_positions WHERE active=1"
    ).fetchall()
    mv: list[tuple[str, str, float]] = []
    for tk, ed, qty, ep in rows:
        px = conn.execute(
            "SELECT close FROM prices WHERE ticker=? ORDER BY date DESC LIMIT 1", (tk,)
        ).fetchone()
        unit = px[0] if px and px[0] else ep
        mv.append((tk, ed, (qty or 0) * unit))
    total = sum(v for _, _, v in mv) or 1.0
    for tk, ed, v in mv:
        conn.execute(
            "UPDATE portfolio_positions SET weight=? WHERE ticker=? AND entry_date=?",
            (v / total, tk, ed),
        )


# ========== BR — XP xlsx ==========

BR_ETF_ADDITIONS = [
    # ticker, name, sector — para inclusão no universe + companies
    ("LFTB11", "iShares Tesouro Selic ETF", "ETF-RF"),
    ("IVVB11", "iShares S&P 500 (BRL)",      "ETF-US"),
]


def _parse_br_date(s: str) -> str | None:
    """'11/04/2024' ou '15/05/2045' → '2024-04-11'. Devolve None se não matchar."""
    if not s:
        return None
    parts = s.strip().split("/")
    if len(parts) != 3:
        return None
    d, m, y = parts
    try:
        return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    except ValueError:
        return None


def _parse_tesouro_maturity(name: str) -> str | None:
    """'NTN-B1 dez/2084' / 'NTNB PRINC ago/2040' → ISO approx '2084-12-15'.
    Tesouro Direto paga no dia 15 do mês."""
    import re
    m = re.search(r'(\w{3,4})/(\d{4})', name.lower())
    if not m:
        return None
    months = {"jan":1, "fev":2, "mar":3, "abr":4, "mai":5, "jun":6,
              "jul":7, "ago":8, "set":9, "out":10, "nov":11, "dez":12}
    mo = months.get(m.group(1)[:3])
    yr = int(m.group(2))
    return f"{yr:04d}-{mo:02d}-15" if mo else None


def _classify_fi(name: str) -> str:
    n = name.upper()
    if n.startswith(("NTN", "LFT", "LTN", "NTNB", "NTN-B", "TESOURO")):
        return "tesouro"
    if n.startswith("DEB"):
        return "debenture"
    if n.startswith("CRA"):
        return "cra"
    if n.startswith("CRI"):
        return "cri"
    if n.startswith("LCA"):
        return "lca"
    if n.startswith("LCI"):
        return "lci"
    return "outro"


def _parse_taxa(taxa_str: str) -> dict:
    """'IPC-A +7,09%' → {indexador: 'IPCA', spread: 0.0709}
    '87,00% CDI'    → {indexador: 'CDI',  cdi_pct: 0.87}
    Retorna dict com chaves indexador, spread_taxa, cdi_pct (qualquer subset)."""
    import re
    if not taxa_str:
        return {}
    s = taxa_str.upper().replace(" ", "")
    out: dict = {}
    # IPCA+X% ou IPC-A+X%
    m = re.match(r'IPC[-]?A\+?([\d,\.]+)%', s)
    if m:
        out["indexador"] = "IPCA"
        out["spread_taxa"] = _parse_brl(m.group(1).replace("%", ""))
        if out["spread_taxa"] is not None:
            out["spread_taxa"] = out["spread_taxa"] / 100.0
        return out
    # X% CDI
    m = re.match(r'([\d,\.]+)%CDI', s)
    if m:
        out["indexador"] = "CDI"
        out["cdi_pct"] = _parse_brl(m.group(1))
        if out["cdi_pct"] is not None:
            out["cdi_pct"] = out["cdi_pct"] / 100.0
        return out
    # CDI+X%
    m = re.match(r'CDI\+?([\d,\.]+)%', s)
    if m:
        out["indexador"] = "CDI"
        out["spread_taxa"] = _parse_brl(m.group(1))
        if out["spread_taxa"] is not None:
            out["spread_taxa"] = out["spread_taxa"] / 100.0
        return out
    # Prefixado
    m = re.match(r'([\d,\.]+)%', s)
    if m:
        out["indexador"] = "PREFIXADO"
        out["spread_taxa"] = _parse_brl(m.group(1))
        if out["spread_taxa"] is not None:
            out["spread_taxa"] = out["spread_taxa"] / 100.0
    return out


def _load_br_fixed_income(xlsx_path: Path) -> list[dict]:
    """Varre secções 'Tesouro Direto' e 'Renda Fixa' do xlsx XP.
    Tesouro (7 cols): nome, pos_atual, alocação, valor_aplicado, qtd, disp, vencimento.
    Renda Fixa (10): nome, pos, alocação, aplicado, aplicado_orig, taxa, aplic_date, venc, qtd, PU.
    """
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    items: list[dict] = []
    section = None
    after_proventos = False
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        c0 = cells[0]
        if c0.startswith("Dividendos") or c0 == "Proventos" or c0 == "Custódia Remunerada":
            after_proventos = True
            section = None
            continue
        if after_proventos:
            continue
        if c0 == "Tesouro Direto":
            section = "tesouro"
            continue
        if c0 == "Renda Fixa":
            section = "rf"
            continue
        if c0 in ("Ações", "Fundos Imobiliários", "Fundos de Investimentos"):
            section = None
            continue
        if section == "tesouro":
            # row: nome, pos, %, aplic, qtd, disp, venc
            kind = _classify_fi(c0)
            if kind != "tesouro":
                continue
            items.append({
                "name": c0, "kind": "tesouro",
                "indexador": "IPCA",  # NTN-B por default (principais são IPCA+)
                "spread_taxa": None,  # Tesouro não expõe taxa na secção simples
                "cdi_pct": None,
                "valor_atual": _parse_brl(cells[1]),
                "valor_aplicado": _parse_brl(cells[3]),
                "quantity": _parse_brl(cells[4]),
                "entry_date": None,
                "maturity_date": _parse_br_date(cells[6]) or _parse_tesouro_maturity(c0),
                "entry_unit_price": None,
            })
        elif section == "rf":
            kind = _classify_fi(c0)
            if kind == "outro":
                continue
            taxa = _parse_taxa(cells[5])
            items.append({
                "name": c0, "kind": kind,
                "indexador": taxa.get("indexador"),
                "spread_taxa": taxa.get("spread_taxa"),
                "cdi_pct": taxa.get("cdi_pct"),
                "valor_atual": _parse_brl(cells[1]),
                "valor_aplicado": _parse_brl(cells[3]),
                "quantity": _parse_brl(cells[8]),
                "entry_date": _parse_br_date(cells[6]),
                "maturity_date": _parse_br_date(cells[7]),
                "entry_unit_price": _parse_brl(cells[9]),
            })
    return items


def _load_br_stocks_fiis(xlsx_path: Path) -> tuple[list[dict], list[dict], dict]:
    """Devolve (acoes, fiis, meta). Só varre as secções 'Ações' e 'Fundos
    Imobiliários'. Header das duas:
      Ações  : ticker, Posição, %, Rent%, PM, Último, Qtd
      FIIs   : ticker, Posição, %, Rent%_prov, Rent%_bruta, PM, Último, Qtd
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    acoes: list[dict] = []
    fiis: list[dict] = []
    meta: dict = {"total_patrimonio": None}

    section = None
    # XP tem secção "Proventos" com sub-blocos "Ações"/"Fundos Imobiliários"
    # que são *dividendos pendentes*, não posições. Após ver Proventos ou
    # Dividendos, bloqueamos as secções de tickers.
    after_proventos = False
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        c0 = cells[0]
        if c0.startswith("R$") and meta["total_patrimonio"] is None:
            meta["total_patrimonio"] = _parse_brl(c0)
        if c0.startswith("Dividendos") or c0 == "Proventos" or c0 == "Custódia Remunerada":
            after_proventos = True
            section = "skip"
            continue
        if not after_proventos:
            if c0 == "Ações":
                section = "acoes"
                continue
            if c0 == "Fundos Imobiliários":
                section = "fiis"
                continue
        if c0 in ("Tesouro Direto", "Fundos de Investimentos", "Renda Fixa"):
            section = "skip"
            continue
        # linhas header das tabelas (começam com "% | ..." ou "0,X% | ...")
        if re.match(r"^\d", c0) and "|" in c0:
            continue

        # linhas de ticker têm tipo "XXXX3" ou "XXXXX11" na primeira coluna
        if section in ("acoes", "fiis") and re.match(r"^[A-Z]{3,6}\d{1,2}$", c0):
            qty_col = 6 if section == "acoes" else 7
            pm_col  = 4 if section == "acoes" else 5
            item = {
                "ticker": c0,
                "posicao": _parse_brl(cells[1]),
                "preco_medio": _parse_brl(cells[pm_col]),
                "ultimo": _parse_brl(cells[pm_col + 1]),
                "quantidade": _parse_brl(cells[qty_col]),
            }
            if section == "acoes":
                acoes.append(item)
            else:
                fiis.append(item)

    return acoes, fiis, meta


def import_br(xlsx_path: Path, dry_run: bool = False) -> dict:
    acoes, fiis, meta = _load_br_stocks_fiis(xlsx_path)
    fi = _load_br_fixed_income(xlsx_path)
    print(f"\n=== BR: Ações ({len(acoes)}) + FIIs ({len(fiis)}) + Renda Fixa ({len(fi)}) ===")
    print(f"  total_patrimonio (XP): R$ {meta['total_patrimonio']:,.2f}")

    if dry_run:
        for a in acoes + fiis:
            q = a.get("quantidade") or 0
            pm = a.get("preco_medio") or 0
            last = a.get("ultimo") or 0
            pos = a.get("posicao") or 0
            print(f"  DRY  {a['ticker']:8s} qty={q:>9.2f}  pm=R${pm:<8.2f}  last=R${last:<8.2f}  pos=R${pos:>12,.2f}")
        print(f"\n  -- Renda Fixa --")
        for f in fi:
            idx = f.get("indexador") or "?"
            spread = f.get("spread_taxa")
            cdi = f.get("cdi_pct")
            taxa_s = f"{idx}+{spread*100:.2f}%" if spread else (f"{cdi*100:.1f}%{idx}" if cdi else idx)
            print(f"  DRY  {f['kind']:<10} {f['name'][:30]:<30} venc={f.get('maturity_date') or '-':<12}  {taxa_s:<15}  atual=R${f.get('valor_atual') or 0:>11,.2f}")
        return {"acoes": acoes, "fiis": fiis, "fi": fi, "meta": meta}

    with sqlite3.connect(DB_BR) as conn:
        ensure_schema(conn)

        # ETFs novos (LFTB11, IVVB11) para entrar em companies
        for tk, nm, sec in BR_ETF_ADDITIONS:
            _upsert_company(conn, tk, nm, "BRL", sec)

        # Limpa imports anteriores da mesma data (permite re-correr)
        conn.execute("DELETE FROM portfolio_positions WHERE entry_date = ?", (TODAY,))

        for a in acoes + fiis:
            if not a["ticker"] or not a["quantidade"] or not a["preco_medio"]:
                continue
            _upsert_position(
                conn, a["ticker"], a["quantidade"], a["preco_medio"],
                TODAY, notes=f"XP import {TODAY}"
            )
            # garante is_holding=1 em companies (idempotente)
            conn.execute(
                "UPDATE companies SET is_holding=1 WHERE ticker=?", (a["ticker"],)
            )

        # Renda fixa: substitui todos os registos XP (idempotente por source='XP')
        conn.execute("DELETE FROM fixed_income_positions WHERE source='XP'")
        now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        for f in fi:
            if not f.get("valor_atual"):
                continue
            conn.execute(
                """INSERT INTO fixed_income_positions
                     (name, kind, indexador, spread_taxa, cdi_pct, entry_date,
                      maturity_date, quantity, entry_unit_price, valor_aplicado,
                      valor_atual, currency, source, fetched_at, notes)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (f["name"], f["kind"], f.get("indexador"),
                 f.get("spread_taxa"), f.get("cdi_pct"), f.get("entry_date"),
                 f.get("maturity_date"), f.get("quantity"),
                 f.get("entry_unit_price"), f.get("valor_aplicado"),
                 f["valor_atual"], "BRL", "XP", now_iso, None),
            )

        _recompute_weights(conn)
        conn.commit()

        total_mv = conn.execute(
            "SELECT sum(quantity * (select close from prices p where p.ticker=pp.ticker order by date desc limit 1)) "
            "FROM portfolio_positions pp WHERE active=1"
        ).fetchone()[0] or 0
        fi_mv = conn.execute(
            "SELECT sum(valor_atual) FROM fixed_income_positions WHERE source='XP'"
        ).fetchone()[0] or 0
        print(f"  persistido: {len(acoes)+len(fiis)} equity + {len(fi)} renda fixa")
        print(f"  MV equity (último close DB): R$ {total_mv:,.2f}")
        print(f"  MV renda fixa (XP snapshot): R$ {fi_mv:,.2f}")
        print(f"  MV BR total                  : R$ {total_mv + fi_mv:,.2f}")

    return {"acoes": acoes, "fiis": fiis, "fi": fi, "meta": meta}


# ========== US — JPM CSV ==========

US_SKIP = {"QACDS", "", None}  # cash sweep + USD cash line


def import_us(csv_path: Path, dry_run: bool = False) -> dict:
    positions: list[dict] = []
    with csv_path.open("r", encoding="utf-8-sig") as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            tk = (row.get("Ticker") or "").strip()
            if not tk or tk in US_SKIP:
                continue
            # JPM escreve "BRKB"; normaliza para "BRK-B"
            if tk == "BRKB":
                tk = "BRK-B"
            qty = _parse_us_number(row.get("Quantity"))
            px = _parse_us_number(row.get("Price"))
            unit_cost = _parse_us_number(row.get("Unit Cost"))
            value = _parse_us_number(row.get("Value"))
            desc = (row.get("Description") or "").strip()
            positions.append({
                "ticker": tk, "quantity": qty, "price": px,
                "unit_cost": unit_cost or px, "value": value,
                "description": desc[:60],
            })

    print(f"\n=== US: {len(positions)} posições ===")
    mv_total = sum(p["value"] or 0 for p in positions)
    print(f"  MV (JPM): $ {mv_total:,.2f}")

    if dry_run:
        for p in positions:
            q = p.get("quantity") or 0
            px = p.get("price") or 0
            uc = p.get("unit_cost") or 0
            val = p.get("value") or 0
            print(f"  DRY  {p['ticker']:7s} qty={q:<10.4f}  px=${px:<9.2f}  uc=${uc:<9.2f}  val=${val:>10,.2f}  [{p['description']}]")
        return {"positions": positions}

    with sqlite3.connect(DB_US) as conn:
        ensure_schema(conn)

        conn.execute("DELETE FROM portfolio_positions WHERE entry_date = ?", (TODAY,))

        for p in positions:
            if not p["quantity"] or not p["unit_cost"]:
                continue
            _upsert_position(
                conn, p["ticker"], p["quantity"], p["unit_cost"],
                TODAY, notes=f"JPM import {TODAY}"
            )
            conn.execute(
                "UPDATE companies SET is_holding=1 WHERE ticker=?", (p["ticker"],)
            )

        _recompute_weights(conn)
        conn.commit()

        total_mv = conn.execute(
            "SELECT sum(quantity * (select close from prices p where p.ticker=pp.ticker order by date desc limit 1)) "
            "FROM portfolio_positions pp WHERE active=1"
        ).fetchone()[0] or 0
        print(f"  persistido: {len(positions)} posições")
        print(f"  MV calculado (último close DB): $ {total_mv:,.2f}")

    return {"positions": positions}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--br", help="Caminho para o xlsx XP PosiçãoDetalhada")
    ap.add_argument("--us", help="Caminho para o csv JPM positions")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.br and not args.us:
        ap.error("precisa --br e/ou --us")

    if args.br:
        import_br(Path(args.br), dry_run=args.dry_run)
    if args.us:
        import_us(Path(args.us), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
